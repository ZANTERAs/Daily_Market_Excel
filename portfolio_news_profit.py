#!/usr/bin/env python3
"""
Portfolio -> Prices, P/L%, Avg Sentiment, and per-ticker News sheets
with dynamic ticker aliases and a TOTAL row.

Key features:
- Reads first sheet of --input Excel:
    Ticker | Buy Price | Buy Date | Shares
- Gets live price (yfinance), computes:
    * Current Price
    * P/L Abs  (dollars)
    * P/L %    (stored as fraction; Excel displays as %)
- Builds Summary sheet with:
    * Avg Sentiment per ticker (mean sentiment over recent news)
    * A TOTAL row aggregating the whole portfolio:
        - Buy Price  -> total cost basis (Σ BuyPrice*Shares)
        - Shares     -> total shares
        - Current Price -> total current value (Σ CurrentPrice*Shares)
        - P/L Abs    -> total $
        - P/L %      -> total % return of the full portfolio
- Builds one "NEWS - {TICKER}" sheet per ticker.
- News comes from your news_harm.py. If a ticker has no coverage,
  we can still fall back to AR-local news if enabled.
- Dynamic aliases:
    * We generate aliases for each ticker using ticker_aliases.build_aliases
      and merge them with aliases.json (user extras) + Argentina defaults.
      Those aliases feed into local AR news queries.
"""

import argparse
import math
from pathlib import Path
import json
import urllib.parse as urlparse

import pandas as pd

# ----------------------------
# Optional imports
# ----------------------------
try:
    import yfinance as yf
except Exception:
    yf = None

try:
    import feedparser
except Exception:
    feedparser = None

try:
    from news_harm import fetch_feeds, map_articles_to_tickers, score_articles
    NEWS_MODULE_OK = True
except Exception:
    NEWS_MODULE_OK = False

try:
    from ticker_aliases import build_aliases as build_dynamic_aliases
    ALIAS_BUILDER_OK = True
except Exception:
    ALIAS_BUILDER_OK = False

# ----------------------------
# Constants / defaults
# ----------------------------
DEFAULT_HEADERS = ["Ticker", "Buy Price", "Buy Date", "Shares"]

DEFAULT_AR_ALIASES = {
    "YPF": ["YPF", "Yacimientos Petrolíferos Fiscales"],
    "PAM": ["Pampa Energía", "Pampa Energia", "Pampa Holding"],
}

GNEWS_BASE = "https://news.google.com/rss/search"


# ----------------------------
# Helpers
# ----------------------------
def ensure_template(path: Path) -> None:
    if path.exists():
        return
    pd.DataFrame(columns=DEFAULT_HEADERS).to_excel(path, sheet_name="Portfolio", index=False)
    print(f"[ok] Created template: {path.name}")

def load_aliases(path: Path) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def dedupe_keep_order(items):
    seen = set()
    out = []
    for it in items:
        if not it:
            continue
        low = str(it).lower().strip()
        if low and low not in seen:
            seen.add(low)
            out.append(str(it).strip())
    return out

def merge_alias_sources(tickers, user_aliases: dict, dyn_aliases: dict) -> dict:
    final_map = {}
    for t in tickers:
        combo = []
        if t in dyn_aliases:
            combo.extend(dyn_aliases[t])
        if t in user_aliases:
            ua = user_aliases[t]
            if isinstance(ua, str):
                combo.append(ua)
            elif isinstance(ua, (list, tuple)):
                combo.extend(ua)
        if t in DEFAULT_AR_ALIASES:
            combo.extend(DEFAULT_AR_ALIASES[t])
        combo.append(t)
        final_map[t] = dedupe_keep_order(combo)
    return final_map

def get_current_price(ticker: str):
    if yf is None:
        return None
    try:
        tk = yf.Ticker(ticker)
        price = None
        # fast_info
        try:
            fi = getattr(tk, "fast_info", None)
            if fi and "lastPrice" in fi:
                price = fi["lastPrice"]
            elif fi and "last_price" in fi:
                price = fi["last_price"]
        except Exception:
            price = None
        # fallback: last close
        if price is None:
            h = tk.history(period="5d", interval="1d", auto_adjust=False)
            if not h.empty:
                price = float(h["Close"].iloc[-1])
        if price is not None and not isinstance(price, (float, int)):
            price = float(price)
        if price is None or (isinstance(price, float) and (math.isnan(price) or price <= 0)):
            return None
        return float(price)
    except Exception:
        return None

# ----------------------------
# Argentina news fetch helpers
# ----------------------------
def build_terms_for_ticker(ticker: str, aliases_map: dict) -> list[str]:
    terms = set([ticker, "Argentina"])
    if ticker in aliases_map:
        for a in aliases_map[ticker]:
            if a:
                terms.add(a)
    return [t for t in terms if t]

def fetch_google_news_ar_for_ticker(
    ticker: str,
    aliases_map: dict,
    days: int = 7,
    max_items: int = 50
) -> pd.DataFrame:
    if feedparser is None:
        return pd.DataFrame(columns=["date","ticker","title","summary","link","source"])

    terms = build_terms_for_ticker(ticker, aliases_map)
    query = "(" + " OR ".join(f'"{t}"' for t in terms) + ") AND (Argentina OR .ar)"
    params = {
        "q": query,
        "hl": "es-419",
        "gl": "AR",
        "ceid": "AR:es-419",
    }
    url = GNEWS_BASE + "?" + urlparse.urlencode(params, doseq=True)

    try:
        feed = feedparser.parse(url)
        rows = []
        for e in feed.entries[:max_items]:
            title = getattr(e, "title", "")
            link  = getattr(e, "link", "")
            published = getattr(e, "published", "") or getattr(e, "updated", "")
            source = ""
            src = getattr(e, "source", None)
            if src is not None:
                source = getattr(src, "title", "") or ""
            rows.append({
                "date": pd.to_datetime(published, errors="coerce"),
                "ticker": ticker,
                "title": title,
                "summary": "",
                "link": link,
                "source": source or "Google News AR",
            })
        df = pd.DataFrame(rows)
        if not df.empty:
            cutoff = pd.Timestamp.today().normalize() - pd.Timedelta(days=days)
            df = df[df["date"] >= cutoff]
        return df
    except Exception:
        return pd.DataFrame(columns=["date","ticker","title","summary","link","source"])

def simple_keyword_sentiment(title: str) -> float:
    t = (title or "").lower()
    neg = [
        "denuncia","demanda","fraude","bancarrota","insolvencia","hackeo","ataque","brecha",
        "cae","baja","recorte","sanción","multa","accidente","explosión","derrame",
        "despidos","downgrade","investigación",
    ]
    pos = [
        "sube","alza","récord","record","mejora","aumenta","aprobación","adquisición",
        "fusión","upgrade",
    ]
    score = 0
    for w in neg:
        if w in t:
            score -= 2
    for w in pos:
        if w in t:
            score += 1
    score = max(-1.0, min(1.0, score/5.0))
    return score

def compute_news_for_tickers(
    tickers,
    backend="vader",
    days=7,
    aliases_map: dict | None = None,
    enable_ar=True
):
    aliases_map = aliases_map or {}
    scored_all = pd.DataFrame(columns=["date","ticker","title","summary","link","source","sentiment"])

    # 1. Try news_harm sources
    if NEWS_MODULE_OK:
        try:
            news = fetch_feeds(tickers)
            if news is not None and not news.empty:
                mapped = map_articles_to_tickers(news, tickers)
                if mapped is not None and not mapped.empty:
                    scored = score_articles(mapped, backend)
                    if scored is not None and not scored.empty:
                        scored_all = scored.copy()
        except Exception:
            pass

    # Ensure cols exist
    for col in ["date","ticker","title","summary","link","source","sentiment"]:
        if col not in scored_all.columns:
            scored_all[col] = pd.Series(dtype="object")

    # 2. Argentina fallback for tickers that still have 0 articles
    if enable_ar:
        if not scored_all.empty:
            counts = scored_all.groupby("ticker").size()
        else:
            import pandas as _pd
            counts = _pd.Series(dtype=int)

        need_fallback = [t for t in tickers if counts.get(t, 0) == 0]
        ar_rows = []
        for t in need_fallback:
            df_ar = fetch_google_news_ar_for_ticker(t, aliases_map, days=days)
            if df_ar is None or df_ar.empty:
                continue

            if NEWS_MODULE_OK:
                try:
                    base_cols = ["date","ticker","title","summary","link","source"]
                    base = df_ar[base_cols].copy()
                    base["summary"] = base.get("summary","")
                    scored_df = score_articles(base, backend)
                    if "sentiment" not in scored_df.columns:
                        scored_df["sentiment"] = [
                            simple_keyword_sentiment(x) for x in scored_df.get("title","")
                        ]
                except Exception:
                    scored_df = df_ar.copy()
                    scored_df["sentiment"] = [
                        simple_keyword_sentiment(x) for x in df_ar["title"]
                    ]
            else:
                scored_df = df_ar.copy()
                scored_df["sentiment"] = [
                    simple_keyword_sentiment(x) for x in df_ar["title"]
                ]

            ar_rows.append(scored_df)

        if ar_rows:
            ar_all = pd.concat(ar_rows, ignore_index=True)
            scored_all = pd.concat([scored_all, ar_all], ignore_index=True)

    # 3. Filter by time window
    if not scored_all.empty:
        cutoff = pd.Timestamp.today().normalize() - pd.Timedelta(days=days)
        scored_all["date"] = pd.to_datetime(scored_all["date"], errors="coerce")
        scored_all = scored_all[scored_all["date"] >= cutoff]

    keep = [c for c in ["date","ticker","title","summary","link","source","sentiment"]
            if c in scored_all.columns]
    return scored_all[keep].copy() if keep else pd.DataFrame(
        columns=["date","ticker","title","summary","link","source","sentiment"]
    )

def write_news_sheet(wb, df_scored: pd.DataFrame, ticker: str):
    from openpyxl.formatting.rule import ColorScaleRule

    sheet_name = f"NEWS - {ticker}"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    ws.append([f"News for {ticker}",
               "(sentiment in [-1,1]; dynamic aliases + AR fallback)"])
    ws.append([])
    ws.append(["Date", "Source", "Title", "Link", "Sentiment"])

    df_t = df_scored[df_scored["ticker"] == ticker].copy()
    if not df_t.empty:
        df_t = df_t.sort_values("date", ascending=False)
        for _, r in df_t.iterrows():
            ws.append([
                pd.to_datetime(r.get("date","")).strftime("%Y-%m-%d")
                if pd.notna(r.get("date","")) else "",
                r.get("source",""),
                r.get("title",""),
                r.get("link",""),
                float(r.get("sentiment", 0.0))
                if pd.notna(r.get("sentiment", None)) else 0.0
            ])

    try:
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 12

        heat = ColorScaleRule(
            start_type='num', start_value=-1,
            mid_type='num', mid_value=0,
            end_type='num', end_value=1,
            start_color='FFC7CE', mid_color='FFEB84', end_color='C6EFCE'
        )
        ws.conditional_formatting.add("E4:E1048576", heat)
    except Exception:
        pass

def build_workbook(
    df_portfolio: pd.DataFrame,
    out_path: Path,
    tickers,
    news_backend="vader",
    news_days=7,
    aliases_map: dict | None = None,
    enable_ar=True
):
    aliases_map = aliases_map or {}

    # ---------------------------------
    # 1) Compute news & avg sentiment
    # ---------------------------------
    df_scored = compute_news_for_tickers(
        tickers,
        backend=news_backend,
        days=news_days,
        aliases_map=aliases_map,
        enable_ar=enable_ar,
    )

    # Per-ticker avg sentiment
    if not df_scored.empty and "sentiment" in df_scored.columns:
        avg_sent = (
            df_scored.groupby("ticker")["sentiment"]
            .mean()
            .rename("Avg Sentiment")
            .to_frame()
        )
    else:
        avg_sent = pd.DataFrame(columns=["Avg Sentiment"])

    # ---------------------------------
    # 2) Build Summary dataframe
    # ---------------------------------
    df_sum = df_portfolio.copy()

    # Per-row dollar P/L and % P/L
    df_sum["P/L Abs"] = (
        (df_sum["Current Price"] - df_sum["Buy Price"]) * df_sum["Shares"]
    )
    df_sum["P/L %"] = (
        (df_sum["Current Price"] - df_sum["Buy Price"]) / df_sum["Buy Price"]
    )  # fraction; Excel shows as %

    # Merge avg sentiment
    df_sum = df_sum.merge(
        avg_sent,
        left_on="Ticker",
        right_index=True,
        how="left"
    )
    if "Avg Sentiment" not in df_sum.columns:
        df_sum["Avg Sentiment"] = pd.NA

    # ---- Compute portfolio totals for the TOTAL row ----
    total_cost_basis = (df_portfolio["Buy Price"] * df_portfolio["Shares"]).sum(skipna=True)
    total_current_value = (df_portfolio["Current Price"] * df_portfolio["Shares"]).sum(skipna=True)
    total_pl_abs = df_sum["P/L Abs"].sum(skipna=True)

    if total_cost_basis and total_cost_basis != 0:
        total_pl_pct = (total_current_value - total_cost_basis) / total_cost_basis
    else:
        total_pl_pct = 0.0

    total_shares = df_portfolio["Shares"].sum(skipna=True)

    # Arrange columns
    ordered_cols = [
        "Ticker",
        "Buy Price",
        "Buy Date",
        "Shares",
        "Current Price",
        "P/L Abs",
        "P/L %",
        "Avg Sentiment",
    ]
    df_sum = df_sum[ordered_cols]

    # Append TOTAL row
    total_row = {
        "Ticker": "TOTAL",
        "Buy Price": total_cost_basis,
        "Buy Date": "",
        "Shares": total_shares,
        "Current Price": total_current_value,
        "P/L Abs": total_pl_abs,
        "P/L %": total_pl_pct,
        "Avg Sentiment": "",
    }
    df_sum_with_total = pd.concat(
        [df_sum, pd.DataFrame([total_row])],
        ignore_index=True
    )

    # ---------------------------------
    # 3) Write Excel
    # ---------------------------------
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        # Original input goes to 'Portfolio'
        df_in = df_portfolio[["Ticker","Buy Price","Buy Date","Shares"]].copy()
        df_in.to_excel(xw, sheet_name="Portfolio", index=False)

        # Summary (incl TOTAL row)
        df_sum_with_total.to_excel(xw, sheet_name="Summary", index=False)

        wb = xw.book
        ws = wb["Summary"]

        # Get the last row (TOTAL row in the sheet)
        last_row = ws.max_row

        # Style columns, conditional formatting, bold TOTAL row
        try:
            from openpyxl.styles import numbers, PatternFill, Font
            from openpyxl.formatting.rule import FormulaRule, ColorScaleRule

            # set column widths
            widths = [12,12,14,10,14,12,10,14]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64+i)].width = w

            # number formats for each row including TOTAL
            for row in range(2, ws.max_row + 1):
                ws[f"E{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Current Price / total current value
                ws[f"F{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # P/L Abs / total P/L Abs
                ws[f"G{row}"].number_format = numbers.FORMAT_PERCENTAGE_00        # P/L %
                ws[f"H{row}"].number_format = "0.00"                              # Avg Sentiment

            # Make TOTAL row bold
            for col_letter in "ABCDEFGH":
                ws[f"{col_letter}{last_row}"].font = Font(bold=True)

            # Also make Buy Price total (column B TOTAL row) show as currency
            ws[f"B{last_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Conditional formatting for P/L % column (G)
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.conditional_formatting.add(
                "G2:G1048576",
                FormulaRule(formula=["G2>0"], fill=green)
            )
            ws.conditional_formatting.add(
                "G2:G1048576",
                FormulaRule(formula=["G2<0"], fill=red)
            )

            # Heatmap for Avg Sentiment (H)
            cscale = ColorScaleRule(
                start_type='num', start_value=-1,
                mid_type='num', mid_value=0,
                end_type='num', end_value=1,
                start_color='FFC7CE', mid_color='FFEB84', end_color='C6EFCE'
            )
            ws.conditional_formatting.add("H2:H1048576", cscale)

            # Move Summary first and freeze panes
            wb.move_sheet(wb["Summary"], offset=-wb.index(wb["Summary"]))
            wb["Portfolio"].freeze_panes = "A2"
            wb["Summary"].freeze_panes   = "A2"
        except Exception:
            pass

        # Per-ticker news sheets
        for t in tickers:
            write_news_sheet(wb, df_scored, t)

        xw._save()

# ----------------------------
# Main CLI
# ----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=str, default="portfolio_input.xlsx", help="Path to input Excel")
    ap.add_argument("--output", type=str, default="portfolio_output.xlsx", help="Path to save output Excel")
    ap.add_argument("--news-backend", type=str, default="vader", choices=["vader","finbert"],
                    help="Sentiment backend used by news_harm.py")
    ap.add_argument("--news-days", type=int, default=7, help="Lookback window for news")
    ap.add_argument("--aliases", type=str, default="aliases.json",
                    help="Optional ticker/company aliases JSON for custom terms")
    ap.add_argument("--ar-news", type=int, default=1,
                    help="Enable AR fallback via Google News (1=yes, 0=no)")
    args = ap.parse_args()

    in_path  = Path(args.input).expanduser().resolve()
    out_path = Path(args.output).expanduser().resolve()
    aliases_path = Path(args.aliases).expanduser().resolve()
    enable_ar = bool(int(args.ar_news))

    ensure_template(in_path)

    df = pd.read_excel(in_path, sheet_name=0)

    # Normalize headers
    rename = {}
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in ["ticker", "symbol"]:
            rename[c] = "Ticker"
        elif cl in ["buy price","buyprice","price","entry price"]:
            rename[c] = "Buy Price"
        elif cl in ["buy date","date","entry date"]:
            rename[c] = "Buy Date"
        elif cl in ["shares","qty","amounts of share","amounts of shares","quantity"]:
            rename[c] = "Shares"
    df = df.rename(columns=rename)

    # Required cols
    for col in ["Ticker","Buy Price","Buy Date","Shares"]:
        if col not in df.columns:
            raise SystemExit(f"Missing required column: {col}")

    # Clean data
    df = df[
        ~df["Ticker"].isna() &
        (df["Ticker"].astype(str).str.strip() != "")
    ].copy()
    df["Ticker"] = df["Ticker"].astype(str).str.upper().str.strip()
    df["Buy Price"] = pd.to_numeric(df["Buy Price"], errors="coerce")
    df["Shares"]    = pd.to_numeric(df["Shares"], errors="coerce").fillna(0).astype(int)

    # Current prices
    prices = []
    for t in df["Ticker"]:
        prices.append(get_current_price(t))
    df["Current Price"] = prices

    # Build alias map
    user_aliases = load_aliases(aliases_path)
    unique_tickers = list(dict.fromkeys(df["Ticker"].tolist()))
    if ALIAS_BUILDER_OK:
        dyn_aliases = build_dynamic_aliases(unique_tickers, extra_aliases=user_aliases)
    else:
        tmp = {}
        for t in unique_tickers:
            al = []
            if t in user_aliases:
                ua = user_aliases[t]
                if isinstance(ua, str):
                    al.append(ua)
                elif isinstance(ua, (list, tuple)):
                    al.extend(ua)
            tmp[t] = dedupe_keep_order(al + [t])
        dyn_aliases = tmp

    final_alias_map = merge_alias_sources(unique_tickers, user_aliases, dyn_aliases)

    # Build workbook (writes Summary, Portfolio, NEWS- sheets)
    build_workbook(
        df_portfolio=df,
        out_path=out_path,
        tickers=unique_tickers,
        news_backend=args.news_backend,
        news_days=args.news_days,
        aliases_map=final_alias_map,
        enable_ar=enable_ar,
    )

    print(f"[ok] Wrote: {out_path.name}")

if __name__ == "__main__":
    main()
